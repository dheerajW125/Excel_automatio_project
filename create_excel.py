import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from datetime import datetime


def Create_Excel(data, output_file):
    wb = Workbook()

    for index, i in enumerate(data):
        if index == 0:
            sheet = wb.active
        else:
            sheet = wb.create_sheet()
        
        sheet.title = i[:31]  

        sheet["A1"] = "M/S . FESTO INDIA PVT LIMITED - FOR THE MONTH OF NOV - 2024"
        sheet.merge_cells("A1:AF1")
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center", indent=1)

        gray_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        sheet["A1"].fill = gray_fill
        for col in range(1, 25):
            sheet.cell(row=2, column=col).fill = gray_fill

        border = Border(
            left=Side(border_style="dotted", color="000000"),
            right=Side(border_style="dotted", color="000000"),
            top=Side(border_style="dotted", color="000000"),
            bottom=Side(border_style="dotted", color="000000")
        )

        sheet.cell(row=1, column=1).border = border
        for col in range(1, 25):
            sheet.cell(row=2, column=col).border = border
        for col in range(1, 33):
            sheet.cell(row=3, column=col).border = border

        sheet["A2"] = "Sl. No."
        sheet["B2"] = "Code"
        sheet["C2"] = "NAME"
        sheet["D2"] = "Designation"
        sheet["E2"] = "DEPT"
        sheet["F2"] = "Days In Month"
        sheet["G2"] = "Days Worked"
        sheet["H2"] = "OT HRS"
        sheet["I2"] = " "
        sheet["J2"] = " "
        sheet["K2"] = "Wage Rate"

        sheet["L2"] = "WAGE RATE"
        sheet.merge_cells("L2:O2")
        sheet["L2"].alignment = Alignment(horizontal="center", vertical="center", indent=1)

        sheet["P2"] = "AMOUNT OF WAGES EARNED"
        sheet.merge_cells("P2:V2")
        sheet["P2"].alignment = Alignment(horizontal="center", vertical="center", indent=1)

        sheet["W2"] = "SALARY DETAILS"
        sheet["X2"] = "BILLING DETAILS"
        sheet.merge_cells("X2:AF2")
        sheet["X2"].alignment = Alignment(horizontal="center", vertical="center", indent=1)

        sheet["F3"] = ""
        sheet["I3"] = "canteen"
        sheet["J3"] = "canteen amt"
        sheet["L3"] = "Basic & D.A"
        sheet["M3"] = "Special Allow / HRA"
        sheet["N3"] = "St Bonus"
        sheet["O3"] = "Leave Wages"
        sheet["P3"] = "Basic & D.A"
        sheet["Q3"] = "Special Allow"
        sheet["R3"] = "St Bonus"
        sheet["S3"] = "Leave Wages"
        sheet["T3"] = "Conveyance"
        sheet["U3"] = "Incentive Amount"
        sheet["V3"] = "OT Amount"
        sheet["W3"] = "GROSS AMOUNT"
        sheet["X3"] = "P . F @ 13%"
        sheet["Y3"] = "ESI @ 3.25%"
        sheet["Z3"] = "LWF"
        sheet["AA3"] = "UNIFORM"
        sheet["AB3"] = "Sub Total"
        sheet["AC3"] = "Service Charges @ 9%"
        sheet["AD3"] = "Total"
        sheet["AE3"] = "CANTEEN"
        sheet["AF3"] = "BILLING AMT"

        for row in sheet.iter_rows():
            max_length = 0
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_height = (max_length + 2) * 1.2 
            sheet.row_dimensions[row[0].row].height = adjusted_height

        wb.save(output_file)

df = pd.read_excel(r"C:\Users\Dheeraj.W\Desktop\MONTH OF APR-24.xlsx", sheet_name='BILLING')
data = []
for i in df['DEPT']:
    i = str(i)
    if ((i != None) and (i.lower() != 'null') and (i.lower()!='nan')):
        data.append(i)
    else:
        break
data = list(set(data))

output_file = "E:\\FORMAT1_OUTPUT\\Wage_Format2" + datetime.now().strftime("%Y-%m-%d %H_%M_%S") + ".xlsx"
Create_Excel(data, output_file)








