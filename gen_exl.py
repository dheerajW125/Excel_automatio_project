import os
import glob
import openpyxl
import pandas as pd
from final import get_row_wise_data_with_headers, base_salary_month_cal
from datetime import datetime

input_filepath = r"C:\Users\Dheeraj.W\Desktop\MONTH OF APR-24.xlsx"
output_filepath = r"E:\FORMAT1_OUTPUT"

files = glob.glob(os.path.join(output_filepath, "*.xlsx"))
most_recent_file = None
most_recent_time = None

for file in files:
    file_time = os.path.getctime(file)
    file_time = datetime.fromtimestamp(file_time)
    
    if most_recent_time is None or file_time > most_recent_time:
        most_recent_time = file_time
        most_recent_file = file

if most_recent_file:
    print(f"Most recent file: {most_recent_file}")
else:
    print("No files found in the directory.")
    exit()

try:
    extracted_data = get_row_wise_data_with_headers(input_filepath)
    # print(extracted_data)
except Exception as e:
    print(f"Error in get_row_wise_data_with_headers: {e}")
    exit()

rows_to_write = []
for row in extracted_data:
    try:
        designation = row["Designation"]
        no_of_days = row["TOTAL PAID DAYS"]
        ot_hrs = row["OT HRS"]
        canteen_deduction = row["Canteen Deduction"]

        if designation in ["SKILLED", "ITI OPERATOR", "ITI"]:
            designation = "Skilled (ITI)"
        elif designation == "DME":
            designation = "DME (2 Yrs)"
        elif designation == "UNSKILLED":
            designation = "Unskilled"

        wage_rate, wage_earned =base_salary_month_cal(designation, no_of_days, ot_hrs, canteen_deduction)
           #wages earned
        base_salary_earned = wage_earned.get("base salary", 0)
        ot_amount_earned = wage_earned.get("OT Amount", 0)
        special_allowance_earned = wage_earned.get("Special allowance", 0)
        st_bonus_earned = wage_earned.get("St Bonus", 0)
        leave_wages_earned = wage_earned.get("Leave Wages", 0)
        pf= wage_earned.get("pf@13", 0)
        uniform = wage_earned.get("Uniform", 0)
        gross_amount = wage_earned.get("Gross Amount", 0)
        service_charges = wage_earned.get("Service charges", 0)
        esi = wage_earned.get("ESI@3.25%", 0)
        sub_total = wage_earned.get("Sub Total", 0)
        total = wage_earned.get("Total", 0)
        canteen = wage_earned.get("canteen", 0)
        billing_amount = wage_earned.get("Billing Amount", 0)
            #wage_rates
        base_salary_rate = wage_rate.get("base salary", 0)
        st_bonus_rate= wage_rate.get("St Bonus", 0)
        ot_amount_rate = wage_rate.get("OT Amount", 0)
        leave_wages_rate = wage_rate.get("Leave Wages", 0)
        special_allowance_rate = wage_rate.get("Special allowance", 0)
        rows_to_write.append({
            "DEPT": row.get("DEPT", ""),
            "Code": row.get("Code", ""),
            "NAME": row.get("Employee Name", ""),
            "Designation": designation,
            "Days In Month": 31, 
            "Days Worked": row["TOTAL PAID DAYS"],
            "OT HRS": row["OT HRS"],
            "canteen":canteen_deduction,
            "canteen amt": canteen,
            "Basic & D.A (rate)":base_salary_rate,
            "Special Allow / HRA (rate)":special_allowance_rate,
            "St Bonus (rate)":st_bonus_rate,
            "Leave Wages (rate)":leave_wages_rate,
            "Basic & D.A (earned)":base_salary_earned,
            "Special Allow / HRA (earned)":special_allowance_earned,
            "St Bonus (earned)":st_bonus_earned,
            "Leave Wages (earned)":leave_wages_earned,
            #"Conveyance"
            #"Incentive Amount"
            "OT Amount": ot_amount_earned,
            "GROSS AMOUNT":gross_amount,
            "P . F @ 13%":pf,
            "ESI @ 3.25%":esi,
            # "LWF":
            "UNIFORM":uniform,
            "Sub Total": sub_total,
            "Service Charges @ 9%": service_charges,
            "Total": total,
            "CANTEEN": canteen,
            "BILLING AMT":billing_amount,
        
        })
    except Exception as e:
        print(f"Error processing row {row}: {e}")

sheet_names = ["OST", "PURCHASE", "QUALITY", "OU-02", "OU-01", "Supervisor", "TEC"]

try:
    workbook = openpyxl.load_workbook(most_recent_file)
    
    for sheet_nm in sheet_names:
        if sheet_nm not in workbook.sheetnames:
            print(f"Sheet '{sheet_nm}' not found in the workbook.")
            continue

        sheet = workbook[sheet_nm]
        header_row = None
        last_row = None 
        
        for row_idx, row in enumerate(sheet.iter_rows(), start=2):#start 2 , becoz header at 1st 2 lines
            if any(cell.value for cell in row):  
                header_row = row_idx
                break

        if header_row is None:
            raise ValueError(f"No header row found in the sheet '{sheet_nm}'")
        last_row = header_row + 1  #writing data after header row
        
        print(f"Header row identified at: {header_row} for sheet '{sheet_nm}'")
        serial_no = 1
        for row_data in rows_to_write:
            if row_data["DEPT"] == sheet_nm:
                last_row += 1  # Increment row for each new data
                sheet[f"A{last_row}"].value = serial_no 
                sheet[f"B{last_row}"].value = row_data["Code"]
                sheet[f"C{last_row}"].value = row_data["NAME"]
                sheet[f"D{last_row}"].value = row_data["Designation"]
                sheet[f"E{last_row}"].value = row_data["DEPT"]
                sheet[f"F{last_row}"].value = row_data["Days In Month"]
                sheet[f"G{last_row}"].value = row_data["Days Worked"]
                sheet[f"H{last_row}"].value = row_data["OT HRS"]
                sheet[f"I{last_row}"].value = row_data["canteen"]
                sheet[f"J{last_row}"].value = row_data["canteen amt"]

                sheet[f"L{last_row}"].value = row_data["Basic & D.A (rate)"]
                sheet[f"M{last_row}"].value = row_data["Special Allow / HRA (rate)"]
                sheet[f"N{last_row}"].value = row_data["St Bonus (rate)"]
                sheet[f"O{last_row}"].value = row_data["Leave Wages (rate)"]
                sheet[f"P{last_row}"].value = row_data["Basic & D.A (earned)"]
                sheet[f"Q{last_row}"].value = row_data["Special Allow / HRA (earned)"]
                sheet[f"R{last_row}"].value = row_data["St Bonus (earned)"]
                sheet[f"S{last_row}"].value = row_data["Leave Wages (earned)"]
                sheet[f"V{last_row}"].value = row_data["OT Amount"]
                sheet[f"W{last_row}"].value = row_data["GROSS AMOUNT"]
                sheet[f"X{last_row}"].value = row_data["P . F @ 13%"]
                sheet[f"Y{last_row}"].value = row_data["ESI @ 3.25%"]
                sheet[f"AA{last_row}"].value = row_data["UNIFORM"]
                sheet[f"AB{last_row}"].value = row_data["Sub Total"]
                sheet[f"AC{last_row}"].value = row_data["Service Charges @ 9%"]
                sheet[f"AD{last_row}"].value = row_data["Total"]
                sheet[f"AE{last_row}"].value = row_data["CANTEEN"]
                sheet[f"AF{last_row}"].value = row_data["BILLING AMT"]

                serial_no += 1 
        
    workbook.save(most_recent_file)
    print(f"Data successfully written to {most_recent_file}")

except Exception as e:
    print(f"Error writing to Excel file: {e}")
